from openpyxl import load_workbook

def order(t):
    s = t[1]
    if s == '学士': return 0
    elif s == '硕士': return 1
    elif s == '博士': return 2
    elif s == '壮士': return 3
    elif s == '圣斗士': return 4
    else: return -1

# 使用 openpyxl 读取数据
wb = load_workbook('学位名单.xlsx', read_only=True)  # read_only 加快读取
ws = wb['Sheet1']
names = [[cell.value for cell in row] for row in ws['B3:C16']]
wb.close()

names.sort(key=order)
for n in names:
    print(n)